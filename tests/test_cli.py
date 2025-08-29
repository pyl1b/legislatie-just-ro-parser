"""Tests for the command line interface."""

import importlib
import json
from pathlib import Path
from types import ModuleType, SimpleNamespace
from typing import Any
from unittest.mock import patch

import yaml  # type: ignore[import-untyped]
from click.testing import CliRunner
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from leropa import cli


def test_convert_outputs_json() -> None:
    """Ensure converting a document id outputs JSON."""

    sample = {"document": {"ver_id": "123"}, "articles": []}
    with patch("leropa.parser.fetch_document", return_value=sample):
        runner = CliRunner()
        result = runner.invoke(cli.cli, ["convert", "123"])

    assert result.exit_code == 0
    assert '"ver_id": "123"' in result.output


def test_convert_writes_json_to_file(tmp_path: Path) -> None:
    """Ensure JSON output is written to a file."""

    sample = {"document": {"ver_id": "123"}, "articles": []}
    out_file = tmp_path / "out.json"

    with patch("leropa.parser.fetch_document", return_value=sample):
        runner = CliRunner()
        result = runner.invoke(
            cli.cli, ["convert", "123", "--output", str(out_file)]
        )

    assert result.exit_code == 0
    assert json.loads(out_file.read_text()) == sample


def test_convert_writes_json_to_directory(tmp_path: Path) -> None:
    """Ensure JSON output is written when a directory is provided."""

    sample = {"document": {"ver_id": "123"}, "articles": []}

    with patch("leropa.parser.fetch_document", return_value=sample):
        runner = CliRunner()
        result = runner.invoke(
            cli.cli, ["convert", "123", "--output", str(tmp_path)]
        )

    out_file = tmp_path / "123.json"
    assert result.exit_code == 0
    assert json.loads(out_file.read_text()) == sample


def test_convert_outputs_yaml() -> None:
    """Ensure YAML output is sent to the console."""

    sample = {"document": {"ver_id": "123"}, "articles": []}
    with patch("leropa.parser.fetch_document", return_value=sample):
        runner = CliRunner()
        result = runner.invoke(cli.cli, ["convert", "123", "--format", "yaml"])

    assert result.exit_code == 0
    assert yaml.safe_load(result.output) == sample


def test_convert_writes_xlsx(tmp_path: Path) -> None:
    """Ensure XLSX output organizes data into PascalCase sheets."""

    sample = {
        "document": {"ver_id": "123"},
        "articles": [
            {
                "article_id": "a1",
                "full_text": "abc",
                "paragraphs": [
                    {
                        "par_id": "p1",
                        "text": "t",
                        "subparagraphs": [],
                        "notes": [],
                    }
                ],
                "notes": [],
            }
        ],
        "books": [
            {
                "book_id": "b1",
                "title": "B1",
                "description": None,
                "titles": [],
                "chapters": [],
                "sections": [],
                "articles": ["a1"],
            }
        ],
    }
    out_file = tmp_path / "out.xlsx"

    with patch("leropa.parser.fetch_document", return_value=sample):
        runner = CliRunner()
        result = runner.invoke(
            cli.cli,
            [
                "convert",
                "123",
                "--format",
                "xlsx",
                "--output",
                str(out_file),
            ],
        )

    assert result.exit_code == 0
    workbook = load_workbook(out_file)
    assert {"Document", "Book", "Article", "Paragraph"} <= set(
        workbook.sheetnames
    )
    book_sheet = workbook["Book"]
    assert book_sheet.cell(row=2, column=1).value == "b1"
    article_sheet = workbook["Article"]
    assert article_sheet.cell(row=2, column=1).value == "a1"
    # Each sheet contains a table named after the sheet.
    assert "Book" in book_sheet.tables
    assert "Article" in article_sheet.tables
    assert "Paragraph" in workbook["Paragraph"].tables


def test_convert_writes_xlsx_to_directory(tmp_path: Path) -> None:
    """Ensure XLSX output is written when a directory is provided."""

    sample = {
        "document": {"ver_id": "123"},
        "articles": [
            {
                "article_id": "a1",
                "full_text": "abc",
                "paragraphs": [],
                "notes": [],
            }
        ],
    }

    with patch("leropa.parser.fetch_document", return_value=sample):
        runner = CliRunner()
        result = runner.invoke(
            cli.cli,
            [
                "convert",
                "123",
                "--format",
                "xlsx",
                "--output",
                str(tmp_path),
            ],
        )

    out_file = tmp_path / "123.xlsx"
    assert result.exit_code == 0
    workbook = load_workbook(out_file)
    assert "Document" in workbook.sheetnames


def test_convert_writes_xlsx_with_nested_values(tmp_path: Path) -> None:
    """Ensure XLSX output serializes nested data structures."""

    sample: dict[str, Any] = {
        "document": {
            "ver_id": "123",
            "versions": [{"ver_id": "2", "date": "2022-01-01"}],
        },
        "articles": [],
    }
    out_file = tmp_path / "out.xlsx"

    with patch("leropa.parser.fetch_document", return_value=sample):
        runner = CliRunner()
        result = runner.invoke(
            cli.cli,
            [
                "convert",
                "123",
                "--format",
                "xlsx",
                "--output",
                str(out_file),
            ],
        )

    assert result.exit_code == 0
    workbook = load_workbook(out_file)
    doc_sheet = workbook["Document"]

    # The versions list is stored as a JSON string in the second column.
    versions_cell = str(doc_sheet.cell(row=2, column=2).value)
    assert json.loads(versions_cell) == sample["document"]["versions"]


def test_convert_generates_missing_ids(tmp_path: Path) -> None:
    """Ensure missing identifiers are generated in the Excel output."""

    sample = {
        "document": {"ver_id": "123"},
        "articles": [
            {
                "article_id": "a1",
                "full_text": "abc",
                "paragraphs": [],
                "notes": [{"text": "note without id"}],
            }
        ],
    }
    out_file = tmp_path / "out.xlsx"

    with patch("leropa.parser.fetch_document", return_value=sample):
        runner = CliRunner()
        result = runner.invoke(
            cli.cli,
            [
                "convert",
                "123",
                "--format",
                "xlsx",
                "--output",
                str(out_file),
            ],
        )

    assert result.exit_code == 0
    workbook = load_workbook(out_file)
    note_sheet = workbook["Note"]
    headers = [cell.value for cell in note_sheet[1]]
    id_col = headers.index("id") + 1
    generated_id = note_sheet.cell(row=2, column=id_col).value
    assert isinstance(generated_id, str) and generated_id.startswith("note_")


def test_convert_wraps_long_and_list_values(tmp_path: Path) -> None:
    """Ensure columns with lists or long text use wrapped cells."""

    sample: dict[str, Any] = {
        "document": {
            "ver_id": "123",
            "summary": "a" * 60,
            "versions": [{"ver_id": "2"}],
        },
        "articles": [
            {
                "article_id": "a1",
                "full_text": "b" * 60,
                "paragraphs": [],
                "notes": [],
            }
        ],
    }
    out_file = tmp_path / "out.xlsx"

    with patch("leropa.parser.fetch_document", return_value=sample):
        runner = CliRunner()
        result = runner.invoke(
            cli.cli,
            [
                "convert",
                "123",
                "--format",
                "xlsx",
                "--output",
                str(out_file),
            ],
        )

    assert result.exit_code == 0
    workbook = load_workbook(out_file)
    doc_sheet = workbook["Document"]
    headers = [cell.value for cell in doc_sheet[1]]
    versions_col = headers.index("versions") + 1
    summary_col = headers.index("summary") + 1
    assert doc_sheet.cell(row=2, column=versions_col).alignment.wrapText
    assert doc_sheet.cell(row=2, column=summary_col).alignment.wrapText
    versions_width = doc_sheet.column_dimensions[
        get_column_letter(versions_col)
    ].width
    summary_width = doc_sheet.column_dimensions[
        get_column_letter(summary_col)
    ].width
    ver_id_col = headers.index("ver_id") + 1
    ver_id_width = doc_sheet.column_dimensions[
        get_column_letter(ver_id_col)
    ].width
    assert versions_width == 50
    assert summary_width == 100
    assert ver_id_width == 12

    article_sheet = workbook["Article"]
    headers = [cell.value for cell in article_sheet[1]]
    full_text_col = headers.index("full_text") + 1
    assert article_sheet.cell(row=2, column=full_text_col).alignment.wrapText
    full_text_width = article_sheet.column_dimensions[
        get_column_letter(full_text_col)
    ].width
    article_id_col = headers.index("article_id") + 1
    article_id_width = article_sheet.column_dimensions[
        get_column_letter(article_id_col)
    ].width
    assert full_text_width == 100
    assert article_id_width == 12


def test_export_md_requires_llm_deps() -> None:
    """Ensure missing LLM deps are reported to the user."""

    runner = CliRunner()
    real_import = importlib.import_module

    def fake_import(name: str, package: str | None = None) -> ModuleType:
        if name.startswith("leropa.llm"):
            raise ModuleNotFoundError
        return real_import(name, package)

    with (
        patch("leropa.cli.importlib.import_module", side_effect=fake_import),
        patch("leropa.cli.click.confirm", return_value=False),
    ):
        result = runner.invoke(cli.cli, ["export-md", "in", "out"])

    assert result.exit_code != 0
    assert "pip install -e .[llm]" in result.output


def test_export_md_invokes_module() -> None:
    """Ensure export command calls the underlying module."""

    called: dict[str, Any] = {}

    def fake_export(
        input_dir: str,
        output_dir: str,
        max_tokens: int,
        overlap_tokens: int,
        title_template: str,
        body_heading: str,
        ext: str,
    ) -> tuple[int, int]:
        called.update(
            {
                "input_dir": input_dir,
                "output_dir": output_dir,
                "max_tokens": max_tokens,
                "overlap_tokens": overlap_tokens,
                "title_template": title_template,
                "body_heading": body_heading,
                "ext": ext,
            }
        )
        return 1, 1

    mod = SimpleNamespace(export_folder=fake_export)
    runner = CliRunner()
    with patch("leropa.cli.importlib.import_module", return_value=mod):
        result = runner.invoke(
            cli.cli,
            [
                "export-md",
                "in_dir",
                "out_dir",
                "--max-tokens",
                "10",
                "--overlap",
                "5",
                "--ext",
                ".txt",
                "--title-template",
                "T",
                "--body-heading",
                "H",
            ],
        )

    assert result.exit_code == 0
    assert called["input_dir"] == "in_dir"
    assert called["output_dir"] == "out_dir"
    assert called["max_tokens"] == 10
    assert called["overlap_tokens"] == 5
    assert called["ext"] == ".txt"
    assert called["title_template"] == "T"
    assert called["body_heading"] == "H"


def test_rag_recreate_invokes_module() -> None:
    """Ensure rag recreate calls into the RAG module."""

    called: dict[str, Any] = {}

    def fake_recreate(collection: str, vector_size: int) -> None:
        called["collection"] = collection
        called["vector_size"] = vector_size

    mod = SimpleNamespace(recreate_collection=fake_recreate)
    runner = CliRunner()
    with patch("leropa.cli.importlib.import_module", return_value=mod):
        result = runner.invoke(cli.cli, ["rag", "recreate", "--dims", "10"])

    assert result.exit_code == 0
    assert called["collection"] == "legal_articles"
    assert called["vector_size"] == 10
