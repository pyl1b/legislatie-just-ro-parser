"""Utilities for exporting document data to Excel workbooks."""

from __future__ import annotations

try:
    import orjson as json
except ImportError:
    import json

from pathlib import Path
from typing import Any, Dict, List
from uuid import uuid4

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.table import (  # type: ignore[import-untyped]
    Table,
    TableStyleInfo,
)

Sheets = Dict[str, List[Dict[str, Any]]]


def _ensure_id(item: Dict[str, Any], prefix: str = "id") -> str:
    """Return an existing id for ``item`` or generate one.

    Args:
        item: Mapping representing a data object.
        prefix: Prefix used when generating a new identifier.

    Returns:
        The unique identifier for ``item``.
    """

    for key, value in item.items():
        if key.endswith("_id") or key == "id":
            return str(value)

    new_id = f"{prefix}_{uuid4().hex}"
    item["id"] = new_id
    return new_id


def _flatten(doc: Dict[str, Any]) -> Sheets:
    """Flatten nested document structure into tabular sheet data.

    Args:
        doc: Parsed document structure.

    Returns:
        Mapping of sheet names to row dictionaries.
    """

    sheets: Sheets = {
        "Document": [],
        "HistoryEntry": [],
        "Book": [],
        "Title": [],
        "Chapter": [],
        "Section": [],
        "Article": [],
        "Paragraph": [],
        "SubParagraph": [],
        "Note": [],
    }

    # Registry of article data keyed by article identifier.
    article_lookup: Dict[str, Dict[str, Any]] = {
        a["article_id"]: a for a in doc.get("articles", [])
    }

    # Map article identifiers to the parent container identifier.
    article_parent: Dict[str, str] = {}

    def process_book(book: Dict[str, Any], parent_id: str) -> str:
        """Process a book structure and its descendants."""

        book_id = _ensure_id(book, "book")
        book["parent_id"] = parent_id

        # Collect identifiers for nested structures.
        title_ids: List[str] = []
        for title in book.get("titles", []):
            title_ids.append(process_title(title, book_id))

        chapter_ids: List[str] = []
        for chapter in book.get("chapters", []):
            chapter_ids.append(process_chapter(chapter, book_id))

        section_ids: List[str] = []
        for section in book.get("sections", []):
            section_ids.append(process_section(section, book_id))

        art_ids = book.get("articles", [])
        for art_id in art_ids:
            article_parent[art_id] = book_id

        book["titles"] = ",".join(title_ids)
        book["chapters"] = ",".join(chapter_ids)
        book["sections"] = ",".join(section_ids)
        book["articles"] = ",".join(art_ids)

        sheets["Book"].append(book)
        return book_id

    def process_title(title: Dict[str, Any], parent_id: str) -> str:
        """Process a title structure and its descendants."""

        title_id = _ensure_id(title, "title")
        title["parent_id"] = parent_id

        chapter_ids: List[str] = []
        for chapter in title.get("chapters", []):
            chapter_ids.append(process_chapter(chapter, title_id))

        section_ids: List[str] = []
        for section in title.get("sections", []):
            section_ids.append(process_section(section, title_id))

        art_ids = title.get("articles", [])
        for art_id in art_ids:
            article_parent[art_id] = title_id

        title["chapters"] = ",".join(chapter_ids)
        title["sections"] = ",".join(section_ids)
        title["articles"] = ",".join(art_ids)

        sheets["Title"].append(title)
        return title_id

    def process_chapter(chapter: Dict[str, Any], parent_id: str) -> str:
        """Process a chapter structure and its descendants."""

        chapter_id = _ensure_id(chapter, "chapter")
        chapter["parent_id"] = parent_id

        section_ids: List[str] = []
        for section in chapter.get("sections", []):
            section_ids.append(process_section(section, chapter_id))

        art_ids = chapter.get("articles", [])
        for art_id in art_ids:
            article_parent[art_id] = chapter_id

        chapter["sections"] = ",".join(section_ids)
        chapter["articles"] = ",".join(art_ids)

        sheets["Chapter"].append(chapter)
        return chapter_id

    def process_section(section: Dict[str, Any], parent_id: str) -> str:
        """Process a section structure and its descendants."""

        section_id = _ensure_id(section, "section")
        section["parent_id"] = parent_id

        subsection_ids: List[str] = []
        for subsection in section.get("subsections", []):
            subsection_ids.append(process_section(subsection, section_id))

        art_ids = section.get("articles", [])
        for art_id in art_ids:
            article_parent[art_id] = section_id

        section["subsections"] = ",".join(subsection_ids)
        section["articles"] = ",".join(art_ids)

        sheets["Section"].append(section)
        return section_id

    def process_paragraph(paragraph: Dict[str, Any], parent_id: str) -> str:
        """Process a paragraph and its substructures."""

        par_id = _ensure_id(paragraph, "par")
        paragraph["parent_id"] = parent_id

        sub_ids: List[str] = []
        for sub in paragraph.get("subparagraphs", []):
            sub_ids.append(process_subparagraph(sub, par_id))

        note_ids: List[str] = []
        for note in paragraph.get("notes", []):
            note_ids.append(process_note(note, par_id))

        paragraph["subparagraphs"] = ",".join(sub_ids)
        paragraph["notes"] = ",".join(note_ids)

        sheets["Paragraph"].append(paragraph)
        return par_id

    def process_subparagraph(
        subparagraph: Dict[str, Any], parent_id: str
    ) -> str:
        """Process a sub-paragraph."""

        sub_id = _ensure_id(subparagraph, "sub")
        subparagraph["parent_id"] = parent_id

        sheets["SubParagraph"].append(subparagraph)
        return sub_id

    def process_note(note: Dict[str, Any], parent_id: str) -> str:
        """Process a note attached to an article or paragraph."""

        note_id = _ensure_id(note, "note")
        note["parent_id"] = parent_id

        sheets["Note"].append(note)
        return note_id

    # Process the document metadata and history entries.
    document = doc.get("document", {})
    doc_id = _ensure_id(document, "doc")

    history_entries = document.pop("history", [])
    history_ids: List[str] = []
    for entry in history_entries:
        entry_id = _ensure_id(entry, "hist")
        entry["parent_id"] = doc_id
        sheets["HistoryEntry"].append(entry)
        history_ids.append(entry_id)

    document["history"] = ",".join(history_ids)

    # Process books and their descendants.
    book_ids: List[str] = []
    for book in doc.get("books", []):
        book_ids.append(process_book(book, doc_id))

    document["books"] = ",".join(book_ids)

    sheets["Document"].append(document)

    # Process all articles, assigning parent references collected earlier.
    for art_id, article in article_lookup.items():
        article_id = _ensure_id(article, "article")
        parent_id = article_parent.get(art_id)
        if parent_id:
            article["parent_id"] = parent_id

        paragraph_ids: List[str] = []
        for paragraph in article.get("paragraphs", []):
            paragraph_ids.append(process_paragraph(paragraph, article_id))

        note_ids: List[str] = []
        for note in article.get("notes", []):
            note_ids.append(process_note(note, article_id))

        article["paragraphs"] = ",".join(paragraph_ids)
        article["notes"] = ",".join(note_ids)

        sheets["Article"].append(article)

    # Drop entries for which no data was recorded.
    return {name: rows for name, rows in sheets.items() if rows}


def write_workbook(doc: Dict[str, Any], path: Path) -> None:
    """Write structured data into an Excel workbook.

    Args:
        doc: Parsed document structure.
        path: Destination file path for the workbook.
    """

    data = _flatten(doc)

    # Create a workbook and remove the default sheet created by openpyxl.
    workbook = Workbook()

    # Remove the default sheet created by openpyxl when present.
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    for sheet_name, rows in data.items():
        ws = workbook.create_sheet(title=sheet_name)

        # Write header row based on dictionary keys.
        headers = list(rows[0].keys())
        ws.append(headers)

        # Track which column indexes require wrapped text and custom widths.
        wrap_columns: set[int] = set()
        list_columns: set[int] = set()
        long_text_columns: set[int] = set()

        for row in rows:
            values: List[Any] = []

            # Serialize complex structures to JSON strings.
            for idx, header in enumerate(headers):
                cell_value = row.get(header)

                # Mark columns with lists or dictionaries for text wrapping and
                # custom width.
                if isinstance(cell_value, (list, dict)):
                    wrap_columns.add(idx)
                    list_columns.add(idx)
                    cell_value = json.dumps(cell_value, ensure_ascii=False)

                # Mark columns containing long text for wrapping and custom
                # width.
                if isinstance(cell_value, str) and len(cell_value) > 50:
                    wrap_columns.add(idx)
                    long_text_columns.add(idx)

                values.append(cell_value)

            ws.append(values)

        # Apply wrap text alignment to the marked columns.
        for col_idx in wrap_columns:
            for col_cells in ws.iter_cols(
                min_col=col_idx + 1,
                max_col=col_idx + 1,
                min_row=1,
                max_row=ws.max_row,
            ):
                for cell in col_cells:
                    cell.alignment = Alignment(wrapText=True)

        # Set column widths based on the contained data type.
        for idx in range(len(headers)):
            col_letter = get_column_letter(idx + 1)
            if idx in list_columns:
                ws.column_dimensions[col_letter].width = 50
            elif idx in long_text_columns:
                ws.column_dimensions[col_letter].width = 100
            else:
                ws.column_dimensions[col_letter].width = 12

        # Determine table range covering the header and all rows.
        end_column = get_column_letter(len(headers))
        end_row = len(rows) + 1
        table = Table(displayName=sheet_name, ref=f"A1:{end_column}{end_row}")

        # Apply a simple table style with row stripes for readability.
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table.tableStyleInfo = style

        ws.add_table(table)

    workbook.save(path)
